import pandas as pd
from openpyxl import load_workbook
from tkinter import filedialog

def main():
    def set_exchangeprice():
        price = -1
        while price < 0 or price > 100:
            try:
                price = float(input("Ingresa el valor de conversión de MXN a USD: "))
                if price < 0 or price > 100:
                    print("Ingresa un valor positivo o realista")
            except ValueError:
                print("Debes ingresar valores válidos")
                
        return price


    def get_tables(df, emptyIdxs: list)->list:
        # Getting the tables and putting them on a list
        tables = []
        # Getting every table on the Excel on a list
        i = 0
        while i < len(emptyIdxs):
            tables.append(df.iloc[emptyIdxs[i] + 1: emptyIdxs[i + 1], :])
            i += 2
            
        return tables
    

    def cleanIndexes(emptyIdxs):
        """
        Function to remove the continuous ranges (ej [300, 301, 302] to [300, 302]) and just stay with the duos.
        TWO POINTERS METHOD
        """
        i, j = 0, 1
        cleanIdxs = []
        
        while j < len(emptyIdxs):
            # If the numbers are a sequence we update j
            if emptyIdxs[j] - emptyIdxs[j - 1] == 1:
                j+=1
                continue
            
            # We put into the new list the new range without the continuous values and update the pointers
            cleanIdxs.extend([emptyIdxs[i], emptyIdxs[j - 1]])
            i = j
            j+=1

        # Return the data clean with the ranges
        return [-1] + cleanIdxs + [emptyIdxs[-1]]
        
    
    def modifyGeneralSheet(df, emptyIdxs, sheet, prices, exchange):
        indexes = []
        for t in get_tables(df, emptyIdxs):
            idxs = t[t.iloc[:, 0].str.startswith('BGM')].index
            indexes.append([idxs[0], idxs[-1]])
        
        for sublist in indexes:
            for i in range(sublist[0], sublist[1] + 1):
                # Calculating the totals to fill up the table (11 is the index of CODIGO DEL PAQUETE)
                productPrice = prices[df.iloc[sublist[0], 11]]
                totalProductPrice = round(productPrice * exchange, 4)
                
                # Writing up the product price
                cell = sheet.cell(row= i + 2, column=14)
                cell.value = productPrice
                
                # Writing up the exchange value
                cell = sheet.cell(row= i + 2, column=15)
                cell.value = exchange
                
                # Writing up the total per product
                cell = sheet.cell(row= i + 2, column=16)
                cell.value = totalProductPrice
                
            # When the loop finishes we write the subtotals
            cell = sheet.cell(row= i + 3, column = 14)
            cell.value = productPrice * (sublist[1] - (sublist[0] - 1))
            
            cell = sheet.cell(row = i + 3, column = 17)
            cell.value = round(totalProductPrice * (sublist[1] - (sublist[0] - 1)), 4)
            
        
    # Prices of the products
    prices = {'010-CCW': 250, '008-CCS': 150, '009-CCP': 150, '001-FPI': 38, '002-FPC': 36, '003-FPP': 80, '004-WWS': 42, '005-INS': 21,
          '006-LSH': 50, '007-LSG': 26, '011-LIS': 35, '012-PCT': 75, '013-SPC': 35, '014-MIG': 10, '015-RPP': 80, '016-SAL': 35,
          '017-IRP': 35, '018-LIS': 75, '019-IRL': 35, '020-PWT': 30, '002-FPC': 36, '001-GAP': 13, '002-GAP': 14, '021-AGT': 25,
          'TFM-001': 80, 'TFM-003': 22, 'TFM-008': 32, 'TFM-009': 21, 'TFM-010': 38, 'TFM-011': 22, 'TFM-012': 495, 'TFM-013': 21,
          'TFM-014': 36, 'GIP-001': 3000, 'CSS-001': 100, '003-GAP': 23, '021-LS6': 85, 'TFM-015': 160, 'TPA-015': 28}
    # ----- GIP ESTA EN MXN CORREGIR ----- #
    
    # Open the file
    file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    
    # Reading the .xlsx file to get the dataframe (the first sheet)
    df = pd.read_excel(file, sheet_name = 'GENERALES')

    # Open the file on openpyxl to write in it
    wb = load_workbook(file)
    # Open the first sheet
    ws = wb['GENERALES']
          
    # Getting the empty indexes  (not clean data) to get all the tables from it
    emptyIdxs = df[df.iloc[:, 0].isna()].index.to_list()

    # Cleaning the indexes
    emptyIdxs = cleanIndexes(emptyIdxs)
    
    # Setting the exchange price
    exchange = set_exchangeprice()
        
    modifyGeneralSheet(df, emptyIdxs, ws, prices, exchange)
    
    # Modify the totals sheet
    totals(df, file, wb, prices, exchange)
        
    print("LISTO")
    
def totals(df, file, wb, prices, exchange):
    def getCardIndexes(list1, list2):
        """
        Method to get the indexes for every card to fill the data and put them together
        """
        indexes = []
        
        for v1, v2 in zip(list1, list2):
            indexes.extend([v1, v2])
            
        return indexes
    
    
    def getCards(indexes, df, headers):
        cards = []
        i = 0
        while i < len(indexes):
            card = df.iloc[indexes[i] + 1: indexes[i + 1] + 1, :]
            card.columns = headers
            
            cards.append(card)
            i+=2
            
        return cards
    
    
    def cleanQuantity(quantity: str) -> int:
        """
        Method to clean the quantities on a str to make them integer
        """
        i = 0
        while i < len(quantity) and quantity[i].isdecimal():
            i += 1
        
        return int(quantity[:i])
    
    
    def getCardsData(cards, prices, headers):
        """
        Method to create a dict with the data needed to fill up the cards.
        """
        res = dict()
        for i, card in enumerate(cards):
            productPriceMX = round(prices[card['CODIGO'].dropna().to_list()[0]] * exchange, 4)
            quantity = cleanQuantity(card[headers[2]][card[headers[2]].str.endswith('M.N.', na = False)].iloc[1])
            subtotal = round(productPriceMX * quantity, 4)
            
            res.update({i: {
                    'quantity': quantity,
                    'productPriceMX': productPriceMX,
                    'subtotalMX': subtotal,
                    'ivaMX': round(subtotal * 0.16, 4),
                    'totalMX': round(subtotal * 1.16, 4),
                    'subtotalUS': round(subtotal / exchange, 2),
                    'ivaUS': round((subtotal / exchange) * 0.16, 2),
                    'totalUS': round((subtotal / exchange) * 1.16, 2)
                }})
        
        return res
    
    
    def formatNumber(num):
        if num % 1 != 0:
            return num
        
        return int(num)
    
    
    def modifyTotalsSheet(df, indexes, headers, exchange, cardData, sheet):
        for i, card in enumerate(getCards(indexes, df, headers)):
            # We create a list with the keys to access to values
            keys = [['subtotalMX', 'subtotalUS'], ['ivaMX', 'ivaUS'], ['totalMX', 'totalUS']]
            keysIdx = 0
            
            # We get the index where the header is what we want
            idx = card.columns.get_loc(headers[2])
            pricesMXIdxs = card[
                (card.iloc[:, idx].str.endswith('M.N.', na = False)) |
                (card.iloc[:, idx].str.endswith('*', na = False))
            ].index.to_list()
            
            # Set the totals in MXN and USD
            sheet.cell(row = pricesMXIdxs[0] + 2, column = idx + 1).value = f'$ {exchange} M.N.'
            sheet.cell(row = pricesMXIdxs[1] + 2, column = idx + 1).value = f'{cardData[i]['quantity']} X {cardData[i]['productPriceMX']:,} Pesos M.N.'
            sheet.cell(row = pricesMXIdxs[1] + 2, column = idx + 2).value = cardData[i]['subtotalMX']
            
            for j in range(len(pricesMXIdxs) - 3, len(pricesMXIdxs)):
                sheet.cell(row = pricesMXIdxs[j] + 2, column = idx + 1).value = f'{cardData[i][keys[keysIdx][0]]:,} Pesos M.N.'
                sheet.cell(row = pricesMXIdxs[j] + 2, column = idx + 2).value = f'{formatNumber(cardData[i][keys[keysIdx][1]]):,} US Dollars'
                keysIdx += 1
                
            sheet.cell(row = card[card.iloc[:, idx] == 'Subtotal'].index[0] + 2, column = idx + 2).value = cardData[i]['subtotalMX']
        
        wb.save(file)
            
            
    ws = wb['TOTALES']
    headers = ['Sample ID', 'CODIGO', 'Sample Form', 'Sample Type']
    df = pd.read_excel(file, sheet_name='TOTALES')
    
    indexes = getCardIndexes(df[df.iloc[:, 1] == headers[1]].index.to_list(), df[df.iloc[:, 1] == 'Total  Amount Due ='].index.to_list())
    
    data = getCardsData(getCards(indexes, df, headers), prices, headers)
    
    modifyTotalsSheet(df, indexes, headers, exchange, data, ws)
    
if __name__ == "__main__":
    main()
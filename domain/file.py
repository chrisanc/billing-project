from openpyxl import load_workbook
import pandas as pd

"""
File class: Used to manage the files in the system
"""
class File:
    def __init__(self, file):
        self.file = file
        self.workbook = self.load_workbook()
        
    def open_sheet(self, sheet_name: str):
        return pd.read_excel(self.file, sheet_name = sheet_name)
        
    def load_workbook(self):
        return load_workbook(self.file)
    
    def load_worksheet(self, sheet_name: str):
        return self.workbook[sheet_name]